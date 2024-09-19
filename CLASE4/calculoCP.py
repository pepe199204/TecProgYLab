import openpyxl
import math

nombreArchivo = "DatosTopografia.xlsx"
libro = openpyxl.load_workbook(nombreArchivo)
hoja = libro.worksheets[0]
x1 = hoja.cell(row=2, column=3).value
y1 = hoja.cell(row=2, column=2).value

fila = 3
while hoja.cell(row=fila, column=1).value != None:
    x2 = hoja.cell(row=fila, column=3).value
    y2 = hoja.cell(row=fila, column=2).value
    
    d = math.sqrt((x1-x2)**2 + (y1-y2)**2)
    a = math.atan((y1-y2)/(x1-x2))
    g = a * 180/math.pi
    
    if y1-y2 > 0:
        if g < 0:
            g = abs(g) + 90
        else:
            g = 270 - g
    else:
        if x1-x2 > 0:
            g = 270 - g
            
    hoja.cell(row=fila, column=7).value = d
    hoja.cell(row=fila, column=8).value = g
    
    fila += 1

libro.save(nombreArchivo)